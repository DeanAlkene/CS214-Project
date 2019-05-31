import openpyxl

# readbook = xlrd.open_workbook(r'.\Project01-Data\TableC-DistanceMatrix.xlsx')
readbook = openpyxl.load_workbook(r'.\Project01-Data\TableA-Orders.xlsx')

sheet = readbook.get_sheet_by_name('Sheet1')

nrows, ncols = sheet.max_row, sheet.max_column

print(nrows,ncols)

l = sheet.cell(1,1).value
print(l)